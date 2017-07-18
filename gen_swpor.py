import pandas as pd
#import matplotlib.pyplot as plt
#from matplotlib.ticker import FuncFormatter
#import pandasql as pdsql
#pysql = lambda q: pdsql.sqldf(q, globals())
import csv
import os

import openpyxl
from openpyxl.utils import coordinate_from_string, column_index_from_string
import util



def gen_swpor(childConfig,configParent):

    config = util.getConfigChild(childConfig)

    inputfile = config['filename']
    inputpath = r'./swpor/' + inputfile
    #path =os.path.basename(os.getcwd())

    #read xlsx and write to csv

    wb = openpyxl.load_workbook(inputpath)
    #sheetnames= wb.get_sheet_names()
    sheet = wb.get_sheet_by_name('SWPOR-Windows 10')

    lst = inputfile.split()
    csvname = lst[0] + '_' + lst[1].split('_')[0][1:].lower() + '.csv'
    csvname_loc = lst[0] + '_' + lst[1].split('_')[0][1:].lower() + '_loc.csv'
    #configParent = util.getConfigParent(parentConfig1[lst[0]])

    print(csvname)

    apps=configParent['apps']
    startCol = config['startCol']
    endCol = config['endCol']
    exclude=config['exclude']
    cyclerow=config['cyclerow']
    platformrow=config['platformrow']


    startCol_loc = config['startCol_loc']
    endCol_loc = config['endCol_loc']
    exclude_loc = config['exclude_loc']
    optioncoderow=config['optioncode_row']

    def getCoords(col1, col2, row1,row2):
        coord1=col1+str(row1)
        coord2=col2+str(row2)
        #print (coord1, coord2)
        return tuple((coord1, coord2))


    def getCol(cooridate):
        xy = coordinate_from_string(cooridate)
        return xy[0]


    # form platform metadata
    def formPlatformMetadata(col1, col2, row1, row2):
        dict1 = dict()
        coords = getCoords(col1,col2,row1,row2)
        tuple(sheet[coords[0]:coords[1]])
        for rowOfCellObjects in sheet[coords[0]:coords[1]]:
            for cellObj in rowOfCellObjects:
                dict1[coordinate_from_string(cellObj.coordinate)[0]] = cellObj.value
                #print(cellObj.coordinate, cellObj.value)
            #print('--- END OF ROW ---')
        return dict1

    dictPlatform = formPlatformMetadata(startCol,endCol,platformrow, platformrow)

    # form App metadata
    def formAppMetadata(col1, col2, row1, row2):
        #apps = config['apps']
        coords = getCoords(col1,col2,row1,row2)
        appdict = dict()

        for rowOfCellObjects in sheet[coords[0]:coords[1]]:
            for cellObj in rowOfCellObjects:
                if cellObj.value not in apps: continue
                # print(cellObj.coordinate, cellObj.value)
                appdict[cellObj.value] = cellObj.coordinate
        return appdict

    dictApp = formAppMetadata('A','A',1, sheet.max_row)

    def getAppRow(appName):
        xy = coordinate_from_string(dictApp[appName])
        return xy[1]

    def getRow(col1,col2,row):
        appCoordStart=col1+str(row)
        appCoordEnd=col2+str(row)
        dictAppPlatform=dict()

        tuple(sheet[appCoordStart:appCoordEnd])
        for rowOfCellObjects in sheet[appCoordStart:appCoordEnd]:
            for cellObj in rowOfCellObjects:
                dictAppPlatform[cellObj.coordinate]=cellObj.value
                #print(cellObj.coordinate, cellObj.value)
                #appdict[cellObj.value] = cellObj.coordinate
        return dictAppPlatform

    def getPlatform(coordinate):
        xy = coordinate_from_string(coordinate)
        col=xy[0]
        return dictPlatform[col]


    # form cycle metadata
    def formCycleMetadata(col1, col2, row1, row2):
        dict1 = dict()
        coords = getCoords(col1,col2,row1,row2)
        lastValue = ''
        tuple(sheet[coords[0]:coords[1]])
        for rowOfCellObjects in sheet[coords[0]:coords[1]]:
            for cellObj in rowOfCellObjects:
                if cellObj.value is not None:
                    lastValue = cellObj.value
                    dict1[coordinate_from_string(cellObj.coordinate)[0]] = cellObj.value
                else:
                    dict1[coordinate_from_string(cellObj.coordinate)[0]] = lastValue
                #print(cellObj.coordinate, cellObj.value, '|', dict1[coordinate_from_string(cellObj.coordinate)[0]])
            #print('--- END OF ROW ---')
        return dict1

    dictCycle = formCycleMetadata(startCol,endCol,cyclerow, cyclerow)








    output = r'./swpor_output/'

    # write to file
    with open(output+csvname, "w") as csv_file:
        writer = csv.writer(csv_file, delimiter=',')
        writer.writerow(['App', 'Cycle', 'Platform'])

        for app in dictApp.items():
            appname= app[0]
            for item in getRow(startCol,endCol,getAppRow(appname)).items():

                if getCol(item[0]) in exclude:
                    continue
                if item[1] is not None:
                    platform = getPlatform(item[0])
                    platform = platform.split(',')[0]
                else: continue
                #print(platform)
                # platform = getPlatform()
                cycle = util.getCycle(item[0],dictCycle,config)

                writer.writerow([appname,cycle,platform])


   # form optioncode metadata
    def formOptionCodeMetadata(col1, col2, row1, row2):
        dict1 = dict()
        coords = getCoords(col1,col2,row1,row2)
        tuple(sheet[coords[0]:coords[1]])
        for rowOfCellObjects in sheet[coords[0]:coords[1]]:
            for cellObj in rowOfCellObjects:
                dict1[coordinate_from_string(cellObj.coordinate)[0]] = cellObj.value
                #print(cellObj.coordinate, cellObj.value)
            #print('--- END OF ROW ---')
        return dict1

    dictOptionCode = formOptionCodeMetadata(startCol_loc,endCol_loc,optioncoderow, optioncoderow)

    def getOptionCode(coordinate):
        xy = coordinate_from_string(coordinate)
        col=xy[0]
        return dictOptionCode[col]

    # write to file
    with open(output+csvname_loc, "w") as csv_file:
        writer = csv.writer(csv_file, delimiter=',')
        writer.writerow(['App', 'OptionCode'])

        for app in dictApp.items():
            appname= app[0]
            for item in getRow(startCol_loc,endCol_loc,getAppRow(appname)).items():
                if getCol(item[0]) in exclude_loc:
                    continue
                if item[1] is not None:
                    optioncode = getOptionCode(item[0])
                else: continue
                #print(platform)
                # platform = getPlatform()
                writer.writerow([appname,optioncode])
